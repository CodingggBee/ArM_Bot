import io
import os
import csv
from pypdf import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook # Lighter than pandas
from fastapi import HTTPException

async def parse_file(filename: str, content: bytes) -> str:
    ext = os.path.splitext(filename)[1].lower()
    try:
        if ext == ".pdf":
            reader = PdfReader(io.BytesIO(content))
            return "\n".join([p.extract_text() for p in reader.pages if p.extract_text()])
        elif ext == ".txt":
            return content.decode("utf-8")
        elif ext == ".csv":
            # Use built-in CSV instead of pandas
            stream = io.StringIO(content.decode("utf-8"))
            return "\n".join([",".join(row) for row in csv.reader(stream)])
        elif ext in [".xlsx", ".xls"]:
            # Use openpyxl instead of pandas
            wb = load_workbook(io.BytesIO(content))
            text = ""
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    text += " ".join([str(cell) for cell in row if cell]) + "\n"
            return text
        elif ext in [".docx", ".doc"]:
            doc = DocxDocument(io.BytesIO(content))
            return "\n".join([p.text for p in doc.paragraphs])
        return ""
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error parsing: {str(e)}")